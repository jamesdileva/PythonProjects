# csv_analyzer_gui.py
# CSV Analyzer with GUI — built with Tkinter and Matplotlib

import pandas as pd
import tkinter as tk
from tkinter import ttk, messagebox, filedialog
import matplotlib.pyplot as plt
import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

THEMES = {
    "light": {
        "bg": "#F0F0F0", "fg": "#1A1A1A",
        "toolbar_bg": "#2E75B6", "summary_bg": "#1F3864",
        "summary_fg": "white", "text_bg": "white",
        "text_fg": "#1A1A1A", "income_bg": "#EAFAF1",
        "expense_bg": "#FDEDEC"
    },
    "dark": {
        "bg": "#1E1E1E", "fg": "#F0F0F0",
        "toolbar_bg": "#1A3A5C", "summary_bg": "#111D2E",
        "summary_fg": "#F0F0F0", "text_bg": "#2D2D2D",
        "text_fg": "#F0F0F0", "income_bg": "#1E3A2F",
        "expense_bg": "#3A1E1E"
    }
}

# ── Data Layer ───────────────────────────────────────────────
def load_csv(filepath):
    df = pd.read_csv(filepath)
    return clean_data(df)

def clean_data(df):
    initial_count = len(df)
    df["amount"] = pd.to_numeric(df["amount"], errors="coerce")
    df = df[df["amount"] >= 0]
    df = df.dropna(subset=["amount"])
    removed = initial_count - len(df)
    return df, removed

# ── Main Application Class ───────────────────────────────────
class CSVAnalyzerApp:
    def __init__(self, root):
        self.root = root
        self.root.title("CSV Analyzer")
        self.root.geometry("1000x700")
        self.root.resizable(True, True)
        self.df = None
        self.filepath = None
        self.theme = "light"
        self._build_ui()

    def _build_ui(self):
        self.root.columnconfigure(0, weight=1)
        self.root.rowconfigure(2, weight=1)

        # ── Top toolbar ──
        toolbar = tk.Frame(self.root, bg="#2E75B6", pady=10)
        toolbar.grid(row=0, column=0, sticky="ew")
        toolbar.columnconfigure(1, weight=1)

        tk.Button(toolbar, text="📂 Load CSV File",
                  command=self._load_file,
                  font=("Arial", 10, "bold"), bg="#27AE60", fg="white",
                  padx=12, pady=4, cursor="hand2").grid(row=0, column=0, padx=10)

        self.lbl_file = tk.Label(toolbar, text="No file loaded",
                                  font=("Arial", 10, "italic"),
                                  bg="#2E75B6", fg="white")
        self.lbl_file.grid(row=0, column=1, sticky="w", padx=8)

        # ── Summary bar ──
        summary_frame = tk.Frame(self.root, bg="#1F3864", pady=8)
        summary_frame.grid(row=1, column=0, sticky="ew")
        summary_frame.columnconfigure([0,1,2,3], weight=1)

        self.lbl_income   = tk.Label(summary_frame, text="Income: —",
            font=("Arial", 12, "bold"), bg="#1F3864", fg="white")
        self.lbl_expenses = tk.Label(summary_frame, text="Expenses: —",
            font=("Arial", 12, "bold"), bg="#1F3864", fg="white")
        self.lbl_balance  = tk.Label(summary_frame, text="Balance: —",
            font=("Arial", 12, "bold"), bg="#1F3864", fg="white")
        self.lbl_rows     = tk.Label(summary_frame, text="Rows: —",
            font=("Arial", 12, "bold"), bg="#1F3864", fg="white")

        self.lbl_income.grid(row=0, column=0)
        self.lbl_expenses.grid(row=0, column=1)
        self.lbl_balance.grid(row=0, column=2)
        self.lbl_rows.grid(row=0, column=3)

        # ── Middle — notebook with tabs ──
        self.notebook = ttk.Notebook(self.root)
        self.notebook.grid(row=2, column=0, sticky="nsew", padx=12, pady=8)

        # Tab 1 — All transactions
        self.tab_data    = tk.Frame(self.notebook)
        # Tab 2 — Category breakdown
        self.tab_cats    = tk.Frame(self.notebook)
        # Tab 3 — Top 5 expenses
        self.tab_top5    = tk.Frame(self.notebook)
        # Tab 4 — Statistics
        self.tab_stats   = tk.Frame(self.notebook)

        self.notebook.add(self.tab_data,  text="  📋 All Transactions  ")
        self.notebook.add(self.tab_cats,  text="  📊 Category Breakdown  ")
        self.notebook.add(self.tab_top5,  text="  🏆 Top 5 Expenses  ")
        self.notebook.add(self.tab_stats, text="  📈 Statistics  ")

        self._build_data_tab()
        self._build_cats_tab()
        self._build_top5_tab()
        self._build_stats_tab()

        # ── Bottom buttons ──
       # ── Bottom buttons ──
        btn_frame = tk.Frame(self.root)
        btn_frame.grid(row=3, column=0, pady=(0, 10))

        btn_style = {"font": ("Arial", 10, "bold"), "width": 18,
                     "pady": 6, "cursor": "hand2"}

        tk.Button(btn_frame, text="📊 Spending Chart",
                  command=self._show_spending_chart,
                  bg="#8E44AD", fg="white", **btn_style).grid(row=0, column=0, padx=6)

        tk.Button(btn_frame, text="📈 Income vs Expenses",
                  command=self._show_income_chart,
                  bg="#16A085", fg="white", **btn_style).grid(row=0, column=1, padx=6)

        tk.Button(btn_frame, text="💾 Export Report",
                  command=self._export_report,
                  bg="#E67E22", fg="white", **btn_style).grid(row=0, column=2, padx=6)

        tk.Button(btn_frame, text="📗 Export to Excel",
                  command=self._export_excel,
                  bg="#217346", fg="white", **btn_style).grid(row=0, column=3, padx=6)

        tk.Button(btn_frame, text="🌙 Toggle Dark Mode",
                  command=self._toggle_theme,
                  bg="#2C3E50", fg="white", **btn_style).grid(row=1, column=0,
                  padx=6, pady=(6,0), columnspan=2)

    def _build_data_tab(self):
        self.tab_data.columnconfigure(0, weight=1)
        self.tab_data.rowconfigure(1, weight=1)

        # Filter bar
        filter_frame = tk.Frame(self.tab_data, pady=6)
        filter_frame.grid(row=0, column=0, sticky="ew", padx=8)

        tk.Label(filter_frame, text="Filter by Category:",
                 font=("Arial", 10)).pack(side="left", padx=(0, 6))

        self.var_filter = tk.StringVar(value="All")
        self.filter_menu = ttk.Combobox(filter_frame,
                                         textvariable=self.var_filter,
                                         state="readonly", width=20)
        self.filter_menu.pack(side="left")
        self.filter_menu.bind("<<ComboboxSelected>>", lambda e: self._apply_filter())

        tk.Button(filter_frame, text="Clear Filter",
                  command=self._clear_filter,
                  font=("Arial", 9), pady=2).pack(side="left", padx=8)

        # Transactions table
        cols = ("type", "category", "amount", "description")
        self.tree_data = ttk.Treeview(self.tab_data, columns=cols,
                                       show="headings", height=20)

        self.tree_data.heading("type",        text="Type")
        self.tree_data.heading("category",    text="Category")
        self.tree_data.heading("amount",      text="Amount")
        self.tree_data.heading("description", text="Description")

        self.tree_data.column("type",        width=90,  anchor="center")
        self.tree_data.column("category",    width=130, anchor="center")
        self.tree_data.column("amount",      width=100, anchor="center")
        self.tree_data.column("description", width=400, anchor="w")

        sb = ttk.Scrollbar(self.tab_data, orient="vertical",
                            command=self.tree_data.yview)
        self.tree_data.configure(yscrollcommand=sb.set)

        self.tree_data.grid(row=1, column=0, sticky="nsew")
        sb.grid(row=1, column=1, sticky="ns")

    def _build_cats_tab(self):
        self.tab_cats.columnconfigure(0, weight=1)
        self.tab_cats.rowconfigure(0, weight=1)

        cols = ("category", "total", "count", "average")
        self.tree_cats = ttk.Treeview(self.tab_cats, columns=cols,
                                       show="headings", height=22)

        self.tree_cats.heading("category", text="Category")
        self.tree_cats.heading("total",    text="Total Spent")
        self.tree_cats.heading("count",    text="Transactions")
        self.tree_cats.heading("average",  text="Average")

        self.tree_cats.column("category", width=180, anchor="center")
        self.tree_cats.column("total",    width=130, anchor="center")
        self.tree_cats.column("count",    width=120, anchor="center")
        self.tree_cats.column("average",  width=130, anchor="center")

        sb = ttk.Scrollbar(self.tab_cats, orient="vertical",
                            command=self.tree_cats.yview)
        self.tree_cats.configure(yscrollcommand=sb.set)

        self.tree_cats.grid(row=0, column=0, sticky="nsew")
        sb.grid(row=0, column=1, sticky="ns")

    def _build_top5_tab(self):
        self.tab_top5.columnconfigure(0, weight=1)
        self.tab_top5.rowconfigure(0, weight=1)

        cols = ("rank", "category", "amount", "description")
        self.tree_top5 = ttk.Treeview(self.tab_top5, columns=cols,
                                       show="headings", height=8)

        self.tree_top5.heading("rank",        text="Rank")
        self.tree_top5.heading("category",    text="Category")
        self.tree_top5.heading("amount",      text="Amount")
        self.tree_top5.heading("description", text="Description")

        self.tree_top5.column("rank",        width=60,  anchor="center")
        self.tree_top5.column("category",    width=160, anchor="center")
        self.tree_top5.column("amount",      width=120, anchor="center")
        self.tree_top5.column("description", width=400, anchor="w")

        self.tree_top5.grid(row=0, column=0, sticky="nsew")

    def _build_stats_tab(self):
        self.tab_stats.columnconfigure(0, weight=1)

        self.stats_text = tk.Text(self.tab_stats, font=("Courier New", 11),
                                   state="disabled", wrap="word",
                                   padx=16, pady=16,
                                   relief="flat")
        self.stats_text.grid(row=0, column=0, sticky="nsew")
        self.tab_stats.rowconfigure(0, weight=1)

    # ── File Loading ─────────────────────────────────────────
    def _load_file(self):
        filepath = filedialog.askopenfilename(
            title="Select a CSV file",
            filetypes=[("CSV files", "*.csv"), ("All files", "*.*")]
        )
        if not filepath:
            return

        try:
            df, removed = load_csv(filepath)
        except Exception as e:
            messagebox.showerror("Load Error",
                f"Could not load file:\n{str(e)}")
            return

        self.df = df
        self.filepath = filepath
        filename = os.path.basename(filepath)
        self.lbl_file.config(text=f"Loaded: {filename}")

        if removed > 0:
            messagebox.showinfo("Data Cleaned",
                f"{removed} invalid row(s) were removed during loading.")

        self._refresh_all()

    # ── Refresh All Views ────────────────────────────────────
    def _refresh_all(self):
        self._update_summary()
        self._update_categories()
        self._populate_filter_menu()
        self._apply_filter()
        self._update_top5()
        self._update_stats()

    def _update_summary(self):
        if self.df is None:
            return
        income   = self.df[self.df["type"] == "income"]["amount"].sum()
        expenses = self.df[self.df["type"] == "expense"]["amount"].sum()
        balance  = income - expenses
        rows     = len(self.df)

        self.lbl_income.config(text=f"Income: ${income:.2f}")
        self.lbl_expenses.config(text=f"Expenses: ${expenses:.2f}")
        self.lbl_balance.config(text=f"Balance: ${balance:.2f}",
            fg="white" if balance >= 0 else "#FFD700")
        self.lbl_rows.config(text=f"Rows: {rows}")

    def _populate_filter_menu(self):
        if self.df is None:
            return
        categories = ["All"] + sorted(self.df["category"].dropna().unique().tolist())
        self.filter_menu["values"] = categories
        self.var_filter.set("All")

    def _apply_filter(self):
        if self.df is None:
            return
        selected = self.var_filter.get()
        filtered = self.df if selected == "All" else \
                   self.df[self.df["category"] == selected]
        self._populate_data_table(filtered)

    def _clear_filter(self):
        self.var_filter.set("All")
        self._apply_filter()

    def _populate_data_table(self, df):
        for row in self.tree_data.get_children():
            self.tree_data.delete(row)
        for _, row in df.iterrows():
            tag = "income" if row["type"] == "income" else "expense"
            self.tree_data.insert("", "end", values=(
                str(row["type"]).capitalize(),
                row["category"],
                f"${row['amount']:.2f}",
                row.get("description", "")
            ), tags=(tag,))
        self.tree_data.tag_configure("income",  background="#EAFAF1")
        self.tree_data.tag_configure("expense", background="#FDEDEC")

    def _update_categories(self):
        for row in self.tree_cats.get_children():
            self.tree_cats.delete(row)
        if self.df is None:
            return
        expenses = self.df[self.df["type"] == "expense"]
        grouped  = expenses.groupby("category")["amount"].agg(
                       ["sum", "count", "mean"]).sort_values("sum", ascending=False)
        for category, row in grouped.iterrows():
            self.tree_cats.insert("", "end", values=(
                category,
                f"${row['sum']:.2f}",
                int(row["count"]),
                f"${row['mean']:.2f}"
            ))

    def _update_top5(self):
        for row in self.tree_top5.get_children():
            self.tree_top5.delete(row)
        if self.df is None:
            return
        expenses = self.df[self.df["type"] == "expense"]
        top5     = expenses.nlargest(5, "amount")
        for rank, (_, row) in enumerate(top5.iterrows(), 1):
            self.tree_top5.insert("", "end", values=(
                f"#{rank}",
                row["category"],
                f"${row['amount']:.2f}",
                row.get("description", "")
            ))

    def _update_stats(self):
        if self.df is None:
            return
        expenses = self.df[self.df["type"] == "expense"]["amount"]
        income   = self.df[self.df["type"] == "income"]["amount"].sum()
        total_exp = expenses.sum()
        balance  = income - total_exp

        lines = [
            "═" * 44,
            "  EXPENSE STATISTICS",
            "═" * 44,
            f"  Highest Expense:   ${expenses.max():.2f}",
            f"  Lowest Expense:    ${expenses.min():.2f}",
            f"  Average Expense:   ${expenses.mean():.2f}",
            f"  Median Expense:    ${expenses.median():.2f}",
            f"  Total Expenses:    ${total_exp:.2f}",
            f"  Total Income:      ${income:.2f}",
            f"  Balance:           ${balance:.2f}",
            f"  Transaction Count: {len(self.df)}",
            f"  Expense Count:     {len(expenses)}",
            "═" * 44,
        ]

        self.stats_text.config(state="normal")
        self.stats_text.delete("1.0", "end")
        self.stats_text.insert("end", "\n".join(lines))
        self.stats_text.config(state="disabled")

    # ── Charts ───────────────────────────────────────────────
    def _show_spending_chart(self):
        if self.df is None:
            messagebox.showinfo("No Data", "Please load a CSV file first.")
            return
        expenses = self.df[self.df["type"] == "expense"]
        if expenses.empty:
            messagebox.showinfo("No Data", "No expenses found in this file.")
            return

        category_totals = expenses.groupby("category")["amount"].sum() \
                                  .sort_values(ascending=False)
        categories = category_totals.index.tolist()
        amounts    = category_totals.values.tolist()
        colors     = ["#2E75B6","#27AE60","#E74C3C","#F39C12",
                      "#8E44AD","#16A085","#D35400","#2C3E50"]

        fig, axes = plt.subplots(1, 2, figsize=(12, 5))
        fig.suptitle("Spending by Category", fontsize=14, fontweight="bold")

        bars = axes[0].bar(categories, amounts,
                           color=colors[:len(categories)],
                           edgecolor="white", linewidth=0.8)
        axes[0].set_title("Expenses by Category")
        axes[0].set_xlabel("Category")
        axes[0].set_ylabel("Amount ($)")
        axes[0].tick_params(axis="x", rotation=30)
        for bar, amount in zip(bars, amounts):
            axes[0].text(bar.get_x() + bar.get_width()/2,
                         bar.get_height() + max(amounts)*0.01,
                         f"${amount:.0f}", ha="center", va="bottom", fontsize=9)

        axes[1].pie(amounts, labels=categories,
                    colors=colors[:len(categories)],
                    autopct="%1.1f%%", startangle=140,
                    wedgeprops={"edgecolor": "white", "linewidth": 1.5})
        axes[1].set_title("Spending Distribution")

        plt.tight_layout()
        plt.show()

    def _show_income_chart(self):
        if self.df is None:
            messagebox.showinfo("No Data", "Please load a CSV file first.")
            return

        income   = self.df[self.df["type"] == "income"]["amount"].sum()
        expenses = self.df[self.df["type"] == "expense"]["amount"].sum()
        balance  = income - expenses

        if income == 0 and expenses == 0:
            messagebox.showinfo("No Data", "No transactions found.")
            return

        fig, axes = plt.subplots(1, 2, figsize=(11, 5))
        fig.suptitle("Income vs Expenses", fontsize=14, fontweight="bold")

        labels = ["Income", "Expenses"]
        values = [income, expenses]
        bar_colors = ["#27AE60", "#E74C3C"]
        bars = axes[0].bar(labels, values, color=bar_colors,
                           edgecolor="white", linewidth=0.8, width=0.5)
        axes[0].set_title("Income vs Expenses")
        axes[0].set_ylabel("Amount ($)")
        for bar, value in zip(bars, values):
            axes[0].text(bar.get_x() + bar.get_width()/2,
                         bar.get_height() + max(values)*0.01,
                         f"${value:.2f}", ha="center", va="bottom",
                         fontsize=10, fontweight="bold")

        axes[1].pie(
            [max(income, 0.01), max(expenses, 0.01)],
            labels=["Income", "Expenses"],
            colors=["#27AE60", "#E74C3C"],
            autopct="%1.1f%%", startangle=90,
            wedgeprops={"edgecolor": "white", "linewidth": 1.5}
        )
        balance_color = "#27AE60" if balance >= 0 else "#E74C3C"
        balance_label = "Surplus" if balance >= 0 else "Deficit"
        axes[1].set_title("Income vs Expenses Split")
        axes[1].text(0, -1.35,
                     f"{balance_label}: ${abs(balance):.2f}",
                     ha="center", fontsize=12,
                     fontweight="bold", color=balance_color)

        plt.tight_layout()
        plt.show()

    # ── Export ───────────────────────────────────────────────
    def _export_report(self):
        if self.df is None:
            messagebox.showinfo("No Data", "Please load a CSV file first.")
            return

        filepath = filedialog.asksaveasfilename(
            defaultextension=".txt",
            filetypes=[("Text files", "*.txt"),
                       ("All files", "*.*")],
            title="Save Report"
        )
        if not filepath:
            return

        income   = self.df[self.df["type"] == "income"]["amount"].sum()
        expenses = self.df[self.df["type"] == "expense"]["amount"].sum()
        balance  = income - expenses
        exp_df   = self.df[self.df["type"] == "expense"]
        breakdown = exp_df.groupby("category")["amount"].sum() \
                          .sort_values(ascending=False)
        top5     = exp_df.nlargest(5, "amount")

        with open(filepath, "w", encoding="utf-8") as f:
            f.write("═" * 44 + "\n")
            f.write("  CSV ANALYZER REPORT\n")
            f.write("═" * 44 + "\n\n")
            f.write(f"  Total Income:   ${income:.2f}\n")
            f.write(f"  Total Expenses: ${expenses:.2f}\n")
            f.write(f"  Balance:        ${balance:.2f}\n\n")
            f.write("  CATEGORY BREAKDOWN\n")
            f.write("  " + "─" * 30 + "\n")
            for cat, total in breakdown.items():
                f.write(f"  {cat:<16} ${total:.2f}\n")
            f.write("\n  TOP 5 EXPENSES\n")
            f.write("  " + "─" * 30 + "\n")
            for rank, (_, row) in enumerate(top5.iterrows(), 1):
                f.write(f"  #{rank} ${row['amount']:.2f} "
                        f"— {row['category']}\n")

        messagebox.showinfo("Exported",
            f"Report saved to:\n{filepath}")
    def _export_excel(self):
        if self.df is None:
            messagebox.showinfo("No Data", "Please load a CSV file first.")
            return

        filepath = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx")],
            title="Save Excel File"
        )
        if not filepath:
            return

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Transactions"

        headers = ["Type", "Category", "Amount", "Description"]
        header_fill = PatternFill("solid", fgColor="1F3864")
        header_font = Font(bold=True, color="FFFFFF", size=11)

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

        income_fill  = PatternFill("solid", fgColor="EAFAF1")
        expense_fill = PatternFill("solid", fgColor="FDEDEC")

        for row, (_, t) in enumerate(self.df.iterrows(), 2):
            ws.cell(row=row, column=1, value=str(t["type"]).capitalize())
            ws.cell(row=row, column=2, value=t["category"])
            ws.cell(row=row, column=3, value=round(float(t["amount"]), 2))
            ws.cell(row=row, column=4, value=str(t.get("description", "")))

            fill = income_fill if t["type"] == "income" else expense_fill
            for col in range(1, 5):
                ws.cell(row=row, column=col).fill = fill

        income   = self.df[self.df["type"] == "income"]["amount"].sum()
        expenses = self.df[self.df["type"] == "expense"]["amount"].sum()
        balance  = income - expenses

        summary_row = len(self.df) + 3
        ws.cell(row=summary_row,     column=3, value="Total Income").font   = Font(bold=True)
        ws.cell(row=summary_row,     column=4, value=round(income, 2)).font = Font(bold=True, color="27AE60")
        ws.cell(row=summary_row + 1, column=3, value="Total Expenses").font = Font(bold=True)
        ws.cell(row=summary_row + 1, column=4, value=round(expenses, 2)).font = Font(bold=True, color="E74C3C")
        ws.cell(row=summary_row + 2, column=3, value="Balance").font = Font(bold=True)
        ws.cell(row=summary_row + 2, column=4, value=round(balance, 2)).font = Font(bold=True)

        ws.column_dimensions["A"].width = 12
        ws.column_dimensions["B"].width = 18
        ws.column_dimensions["C"].width = 12
        ws.column_dimensions["D"].width = 35

        wb.save(filepath)
        messagebox.showinfo("Exported",
            f"Transactions exported to:\n{filepath}")

    def _toggle_theme(self):
        self.theme = "dark" if self.theme == "light" else "light"
        t = THEMES[self.theme]
        self.root.configure(bg=t["bg"])
        self._apply_theme_recursive(self.root, t)

        self.stats_text.config(
            bg=t["text_bg"],
            fg=t["text_fg"]
        )

        self.tree_data.tag_configure("income",  background=t["income_bg"])
        self.tree_data.tag_configure("expense", background=t["expense_bg"])

    def _apply_theme_recursive(self, widget, t):
        widget_type = widget.winfo_class()
        try:
            if widget_type in ("Frame", "Label"):
                widget.configure(bg=t["bg"], fg=t["fg"])
            elif widget_type == "Button":
                pass
        except tk.TclError:
            pass
        for child in widget.winfo_children():
            self._apply_theme_recursive(child, t)

# ── Entry Point ──────────────────────────────────────────────
if __name__ == "__main__":
    root = tk.Tk()
    app = CSVAnalyzerApp(root)
    root.mainloop()
