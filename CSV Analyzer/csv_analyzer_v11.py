# csv_analyzer_v11.py
# CSV Analyzer v1.1 — CustomTkinter modern UI
# Slate and Amber color scheme — analyst tool layout

import pandas as pd
import os
import customtkinter as ctk
from tkinter import ttk, messagebox, filedialog
import matplotlib.pyplot as plt
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import datetime

# ── CustomTkinter Setup ──────────────────────────────────────
ctk.set_appearance_mode("light")
ctk.set_default_color_theme("blue")

# ── Color Palette ────────────────────────────────────────────
COLORS = {
    "slate":        "#2B3A4A",
    "slate_light":  "#3D5166",
    "slate_dark":   "#1A2530",
    "amber":        "#F39C12",
    "amber_dark":   "#D68910",
    "white":        "#FFFFFF",
    "off_white":    "#F8F9FA",
    "grey_light":   "#E9ECEF",
    "grey":         "#6C757D",
    "text_dark":    "#1A1A2E",
    "text_light":   "#F8F9FA",
    "income_bg":    "#D5F5E3",
    "expense_bg":   "#FADBD8",
    "green":        "#27AE60",
    "red":          "#E74C3C",
    "dark_bg":      "#1A2530",
    "dark_panel":   "#2B3A4A",
    "dark_surface": "#3D5166",
}

# ── Data Layer ───────────────────────────────────────────────
def load_csv(filepath):
    df = pd.read_csv(filepath)
    return clean_data(df)

def clean_data(df):
    initial_count = len(df)
    df["amount"] = pd.to_numeric(df["amount"], errors="coerce")
    df = df[df["amount"] >= 0]
    df = df.dropna(subset=["amount"])
    removed = initial_count - len(df)
    return df, removed

# ── Main Application ─────────────────────────────────────────
class CSVAnalyzerV11(ctk.CTk):
    def __init__(self):
        super().__init__()

        self.title("CSV Analyzer")
        self.geometry("1200x720")
        self.minsize(1000, 640)

        self.df        = None
        self.filepath  = None
        self.dark_mode = False

        self._apply_colors()
        self._build_ui()

    def _apply_colors(self):
        self.C = COLORS
        if self.dark_mode:
            ctk.set_appearance_mode("dark")
            self.bg          = self.C["dark_bg"]
            self.panel_bg    = self.C["dark_panel"]
            self.surface_bg  = self.C["dark_surface"]
            self.text_color  = self.C["text_light"]
            self.sub_color   = "#AABBCC"  # was #8899AA — brighter
            self.income_row  = "#1A3A2A"
            self.expense_row = "#3A1A1A"
        else:
            ctk.set_appearance_mode("light")
            self.bg          = self.C["off_white"]
            self.panel_bg    = self.C["white"]
            self.surface_bg  = self.C["grey_light"]
            self.text_color  = self.C["text_dark"]
            self.sub_color   = self.C["grey"]
            self.income_row  = self.C["income_bg"]
            self.expense_row = self.C["expense_bg"]

    # ── UI Construction ──────────────────────────────────────
    def _build_ui(self):
        self.configure(fg_color=self.bg)
        self.columnconfigure(0, weight=1)
        self.rowconfigure(1, weight=1)

        self._build_toolbar()
        self._build_main_area()
        self._build_bottom_buttons()

    def _build_toolbar(self):
        self.toolbar = ctk.CTkFrame(self,
                                     fg_color=self.C["slate"],
                                     corner_radius=0,
                                     height=64)
        self.toolbar.grid(row=0, column=0, sticky="ew")
        self.toolbar.grid_propagate(False)
        self.toolbar.columnconfigure(3, weight=1)

        # Title
        ctk.CTkLabel(self.toolbar,
                      text="📊  CSV Analyzer",
                      font=ctk.CTkFont("Arial", 18, "bold"),
                      text_color=self.C["white"]).grid(
                      row=0, column=0, padx=(20,4),
                      pady=16, sticky="w")

        ctk.CTkLabel(self.toolbar,
                      text="analyst tool",
                      font=ctk.CTkFont("Arial", 10),
                      text_color="#8899AA").grid(
                      row=0, column=1, padx=(0,24),
                      pady=16, sticky="w")

        # Load button
        ctk.CTkButton(self.toolbar,
                       text="📂  Load CSV",
                       command=self._load_file,
                       fg_color=self.C["amber"],
                       hover_color=self.C["amber_dark"],
                       text_color=self.C["white"],
                       font=ctk.CTkFont("Arial", 11, "bold"),
                       height=36,
                       corner_radius=8,
                       width=130).grid(
                       row=0, column=2, padx=(0,12), pady=14)

        # Filter
        self.var_filter = ctk.StringVar(value="All Categories")
        self.filter_menu = ctk.CTkOptionMenu(
            self.toolbar,
            variable=self.var_filter,
            values=["All Categories"],
            command=self._apply_filter,
            fg_color=self.C["slate_light"],
            button_color=self.C["amber"],
            button_hover_color=self.C["amber_dark"],
            text_color=self.C["white"],
            font=ctk.CTkFont("Arial", 11),
            width=180,
            height=36)
        self.filter_menu.grid(row=0, column=4,
                               padx=(0,12), pady=14)

        # Status label
        self.lbl_status = ctk.CTkLabel(
            self.toolbar,
            text="Load a CSV file to begin",
            font=ctk.CTkFont("Arial", 10),
            text_color="#8899AA")
        self.lbl_status.grid(row=0, column=5,
                              padx=(0,12), pady=14,
                              sticky="w")

        # Dark mode
        self.btn_theme = ctk.CTkButton(
            self.toolbar,
            text="🌙",
            command=self._toggle_theme,
            fg_color=self.C["slate_light"],
            hover_color=self.C["slate_dark"],
            text_color=self.C["white"],
            font=ctk.CTkFont("Arial", 14),
            width=40,
            height=36,
            corner_radius=8)
        self.btn_theme.grid(row=0, column=6,
                             padx=(0,20), pady=14)

    def _build_main_area(self):
        main = ctk.CTkFrame(self, fg_color=self.bg,
                             corner_radius=0)
        main.grid(row=1, column=0, sticky="nsew",
                  padx=16, pady=(12,8))
        main.columnconfigure(0, weight=1)
        main.columnconfigure(1, minsize=240)
        main.rowconfigure(0, weight=1)

        self._build_data_table(main)
        self._build_stats_panel(main)

    def _build_data_table(self, parent):
        table_frame = ctk.CTkFrame(parent,
                                    fg_color=self.panel_bg,
                                    corner_radius=12)
        table_frame.grid(row=0, column=0, sticky="nsew",
                          padx=(0,12))
        table_frame.columnconfigure(0, weight=1)
        table_frame.rowconfigure(1, weight=1)

        ctk.CTkLabel(table_frame,
                      text="Transactions",
                      font=ctk.CTkFont("Arial", 14, "bold"),
                      text_color=self.text_color).grid(
                      row=0, column=0, sticky="w",
                      padx=16, pady=(14,8))

        style = ttk.Style()
        style.theme_use("clam")
        style.configure("Analyst.Treeview",
                         background=self.panel_bg,
                         foreground=self.text_color,
                         fieldbackground=self.panel_bg,
                         rowheight=30,
                         font=("Arial", 10))
        style.configure("Analyst.Treeview.Heading",
                         background=self.C["slate"],
                         foreground=self.C["white"],
                         font=("Arial", 10, "bold"),
                         relief="flat")
        style.map("Analyst.Treeview",
                   background=[("selected",
                                 self.C["slate_light"])])

        cols = ("type", "category", "amount", "description")
        self.tree = ttk.Treeview(table_frame,
                                  columns=cols,
                                  show="headings",
                                  style="Analyst.Treeview",
                                  height=20)

        self.tree.heading("type",        text="Type")
        self.tree.heading("category",    text="Category")
        self.tree.heading("amount",      text="Amount")
        self.tree.heading("description", text="Description")

        self.tree.column("type",        width=90,  anchor="center")
        self.tree.column("category",    width=130, anchor="center")
        self.tree.column("amount",      width=100, anchor="center")
        self.tree.column("description", width=320, anchor="w")

        sb = ttk.Scrollbar(table_frame, orient="vertical",
                            command=self.tree.yview)
        self.tree.configure(yscrollcommand=sb.set)
        self.tree.grid(row=1, column=0, sticky="nsew",
                        padx=(12,0), pady=(0,12))
        sb.grid(row=1, column=1, sticky="ns", pady=(0,12))

    def _build_stats_panel(self, parent):
        self.stats_panel = ctk.CTkScrollableFrame(
            parent,
            fg_color=self.panel_bg,
            corner_radius=12,
            scrollbar_button_color=self.C["slate_light"],
            scrollbar_button_hover_color=self.C["amber"],
            width=240)
        self.stats_panel.grid(row=0, column=1,
                               sticky="nsew")
        self.stats_panel.columnconfigure(0, weight=1)

        self._build_summary_section()
        self._build_categories_section()
        self._build_stats_section()

    def _build_summary_section(self):
        ctk.CTkLabel(self.stats_panel,
                      text="SUMMARY",
                      font=ctk.CTkFont("Arial", 10, "bold"),
                      text_color=self.C["amber"]).pack(
                      padx=16, pady=(16,8), anchor="w")

        self.lbl_income = self._stat_row(
            "Income", "$—", self.C["green"])
        self.lbl_expenses = self._stat_row(
            "Expenses", "$—", self.C["red"])
        self.lbl_balance = self._stat_row(
            "Balance", "$—", self.C["amber"])
        self.lbl_rows = self._stat_row(
            "Rows", "—", self.C["slate"])

        self._divider()

    def _build_categories_section(self):
        ctk.CTkLabel(self.stats_panel,
                      text="TOP CATEGORIES",
                      font=ctk.CTkFont("Arial", 10, "bold"),
                      text_color=self.C["amber"]).pack(
                      padx=16, pady=(8,8), anchor="w")

        self.cats_frame = ctk.CTkFrame(
            self.stats_panel,
            fg_color="transparent")
        self.cats_frame.pack(fill="x", padx=16)

        self._divider()

    def _build_stats_section(self):
        ctk.CTkLabel(self.stats_panel,
                      text="STATISTICS",
                      font=ctk.CTkFont("Arial", 10, "bold"),
                      text_color=self.C["amber"]).pack(
                      padx=16, pady=(8,8), anchor="w")

        self.lbl_highest = self._stat_row(
            "Highest", "$—", self.C["red"])
        self.lbl_lowest  = self._stat_row(
            "Lowest",  "$—", self.C["green"])
        self.lbl_average = self._stat_row(
            "Average", "$—", self.C["slate"])
        self.lbl_median  = self._stat_row(
            "Median",  "$—", self.C["slate"])

    def _stat_row(self, label, value, color):
        row = ctk.CTkFrame(self.stats_panel,
                            fg_color="transparent")
        row.pack(fill="x", padx=16, pady=(0,6))
        row.columnconfigure(1, weight=1)

        self._sub_labels = getattr(self, "_sub_labels", [])

        lbl_name = ctk.CTkLabel(row, text=label,
                      font=ctk.CTkFont("Arial", 10),
                      text_color=self.sub_color)
        lbl_name.grid(row=0, column=0, sticky="w")
        self._sub_labels.append(lbl_name)

        lbl = ctk.CTkLabel(row, text=value,
                            font=ctk.CTkFont("Arial", 11, "bold"),
                            text_color=color)
        lbl.grid(row=0, column=1, sticky="e")
        return lbl

    def _divider(self):
        ctk.CTkFrame(self.stats_panel,
                      height=1,
                      fg_color=self.surface_bg).pack(
                      fill="x", padx=16, pady=8)

    def _build_bottom_buttons(self):
        btn_frame = ctk.CTkFrame(self,
                                  fg_color="transparent")
        btn_frame.grid(row=2, column=0, sticky="ew",
                        padx=16, pady=(0,16))

        btn_cfg = {
            "font": ctk.CTkFont("Arial", 11),
            "height": 36,
            "corner_radius": 8
        }

        ctk.CTkButton(btn_frame,
                       text="📊 Spending Chart",
                       command=self._show_spending_chart,
                       fg_color="#8E44AD",
                       hover_color="#7D3C98",
                       text_color=self.C["white"],
                       **btn_cfg).pack(side="left", padx=(0,8))

        ctk.CTkButton(btn_frame,
                       text="📈 Income vs Expenses",
                       command=self._show_income_chart,
                       fg_color="#16A085",
                       hover_color="#138D75",
                       text_color=self.C["white"],
                       **btn_cfg).pack(side="left", padx=(0,8))

        ctk.CTkButton(btn_frame,
                       text="💾 Export Excel",
                       command=self._export_excel,
                       fg_color="#217346",
                       hover_color="#1A5C38",
                       text_color=self.C["white"],
                       **btn_cfg).pack(side="left", padx=(0,8))

        ctk.CTkButton(btn_frame,
                       text="📄 Export Report",
                       command=self._export_report,
                       fg_color=self.C["amber"],
                       hover_color=self.C["amber_dark"],
                       text_color=self.C["white"],
                       **btn_cfg).pack(side="left")

    # ── Status Helper ────────────────────────────────────────
    def _set_status(self, message, color="#8899AA"):
        self.lbl_status.configure(text=message,
                                   text_color=color)
        self.update_idletasks()

    # ── Load File ────────────────────────────────────────────
    def _load_file(self):
        filepath = filedialog.askopenfilename(
            title="Select a CSV file",
            filetypes=[("CSV files", "*.csv"),
                       ("All files", "*.*")]
        )
        if not filepath:
            return

        try:
            df, removed = load_csv(filepath)
        except Exception as e:
            messagebox.showerror("Load Error",
                f"Could not load file:\n{str(e)}")
            return

        self.df       = df
        self.filepath = filepath
        filename      = os.path.basename(filepath)

        if removed > 0:
            messagebox.showinfo("Data Cleaned",
                f"{removed} invalid row(s) removed.")

        self._set_status(
            f"✓ {filename} — {len(df)} rows loaded",
            self.C["amber"])
        self._refresh_all()

    # ── Refresh All ──────────────────────────────────────────
    def _refresh_all(self):
        self._update_table()
        self._update_summary()
        self._update_categories()
        self._update_stats()
        self._update_filter_menu()

    def _update_table(self, df=None):
        data = df if df is not None else self.df
        if data is None:
            return

        for row in self.tree.get_children():
            self.tree.delete(row)

        for _, row in data.iterrows():
            tag = "income" if row["type"] == "income" \
                  else "expense"
            self.tree.insert("", "end", values=(
                str(row["type"]).capitalize(),
                row["category"],
                f"${row['amount']:.2f}",
                str(row.get("description", ""))
            ), tags=(tag,))

        self.tree.tag_configure("income",
                                  background=self.income_row)
        self.tree.tag_configure("expense",
                                  background=self.expense_row)

    def _update_summary(self):
        if self.df is None:
            return
        income   = self.df[self.df["type"] == "income"]["amount"].sum()
        expenses = self.df[self.df["type"] == "expense"]["amount"].sum()
        balance  = income - expenses

        self.lbl_income.configure(text=f"${income:.2f}")
        self.lbl_expenses.configure(text=f"${expenses:.2f}")
        self.lbl_balance.configure(
            text=f"${balance:.2f}",
            text_color=self.C["green"] if balance >= 0
                       else self.C["red"])
        self.lbl_rows.configure(text=str(len(self.df)))

    def _update_categories(self):
        for widget in self.cats_frame.winfo_children():
            widget.destroy()

        if self.df is None:
            return

        expenses = self.df[self.df["type"] == "expense"]
        if expenses.empty:
            return

        top5 = expenses.groupby("category")["amount"] \
                        .sum().sort_values(ascending=False).head(5)

        for category, total in top5.items():
            row = ctk.CTkFrame(self.cats_frame,
                                fg_color="transparent")
            row.pack(fill="x", pady=(0,4))
            row.columnconfigure(1, weight=1)

            ctk.CTkLabel(row,
                          text=category,
                          font=ctk.CTkFont("Arial", 10),
                          text_color=self.text_color).grid(
                          row=0, column=0, sticky="w")

            ctk.CTkLabel(row,
                          text=f"${total:.0f}",
                          font=ctk.CTkFont("Arial", 10, "bold"),
                          text_color=self.C["amber"]).grid(
                          row=0, column=1, sticky="e")

    def _update_stats(self):
        if self.df is None:
            return
        expenses = self.df[self.df["type"] == "expense"]["amount"]
        if expenses.empty:
            return

        self.lbl_highest.configure(
            text=f"${expenses.max():.2f}")
        self.lbl_lowest.configure(
            text=f"${expenses.min():.2f}")
        self.lbl_average.configure(
            text=f"${expenses.mean():.2f}")
        self.lbl_median.configure(
            text=f"${expenses.median():.2f}")

    def _update_filter_menu(self):
        if self.df is None:
            return
        categories = ["All Categories"] + sorted(
            self.df["category"].dropna().unique().tolist())
        self.filter_menu.configure(values=categories)
        self.var_filter.set("All Categories")

    def _apply_filter(self, selected=None):
        if self.df is None:
            return
        sel = selected or self.var_filter.get()
        filtered = self.df if sel == "All Categories" \
                   else self.df[self.df["category"] == sel]
        self._update_table(filtered)

    # ── Charts ───────────────────────────────────────────────
    def _show_spending_chart(self):
        if self.df is None:
            messagebox.showinfo("No Data",
                "Please load a CSV file first.")
            return

        expenses = self.df[self.df["type"] == "expense"]
        if expenses.empty:
            messagebox.showinfo("No Data",
                "No expenses found.")
            return

        cat_totals = expenses.groupby("category")["amount"] \
                             .sum().sort_values(ascending=False)
        categories = cat_totals.index.tolist()
        amounts    = cat_totals.values.tolist()
        colors     = ["#F39C12","#E74C3C","#2ECC71","#8E44AD",
                      "#16A085","#2E75B6","#D35400","#2C3E50"]

        fig, axes = plt.subplots(1, 2, figsize=(12, 5))
        fig.patch.set_facecolor("#1A2530")
        for ax in axes:
            ax.set_facecolor("#2B3A4A")
            ax.tick_params(colors="white")
            ax.xaxis.label.set_color("white")
            ax.yaxis.label.set_color("white")
            ax.title.set_color("white")
            for spine in ax.spines.values():
                spine.set_edgecolor("#3D5166")

        fig.suptitle("Spending by Category",
                     fontsize=14, fontweight="bold",
                     color="white")

        bars = axes[0].bar(categories, amounts,
                           color=colors[:len(categories)],
                           edgecolor="#1A2530",
                           linewidth=0.8)
        axes[0].set_title("Expenses by Category")
        axes[0].set_xlabel("Category")
        axes[0].set_ylabel("Amount ($)")
        axes[0].tick_params(axis="x", rotation=30)

        for bar, amount in zip(bars, amounts):
            axes[0].text(
                bar.get_x() + bar.get_width()/2,
                bar.get_height() + max(amounts)*0.01,
                f"${amount:.0f}",
                ha="center", va="bottom",
                fontsize=9, color="white")

        axes[1].pie(amounts, labels=categories,
                    colors=colors[:len(categories)],
                    autopct="%1.1f%%", startangle=140,
                    wedgeprops={"edgecolor": "#1A2530",
                                "linewidth": 1.5},
                    textprops={"color": "white"})
        axes[1].set_title("Spending Distribution")

        plt.tight_layout()
        plt.show()

    def _show_income_chart(self):
        if self.df is None:
            messagebox.showinfo("No Data",
                "Please load a CSV file first.")
            return

        income   = self.df[self.df["type"] == "income"]["amount"].sum()
        expenses = self.df[self.df["type"] == "expense"]["amount"].sum()
        balance  = income - expenses

        if income == 0 and expenses == 0:
            messagebox.showinfo("No Data",
                "No transactions found.")
            return

        fig, axes = plt.subplots(1, 2, figsize=(11, 5))
        fig.patch.set_facecolor("#1A2530")
        for ax in axes:
            ax.set_facecolor("#2B3A4A")
            ax.tick_params(colors="white")
            for spine in ax.spines.values():
                spine.set_edgecolor("#3D5166")

        fig.suptitle("Income vs Expenses",
                     fontsize=14, fontweight="bold",
                     color="white")

        labels     = ["Income", "Expenses"]
        values     = [income, expenses]
        bar_colors = ["#2ECC71", "#E74C3C"]

        bars = axes[0].bar(labels, values,
                           color=bar_colors,
                           edgecolor="#1A2530",
                           linewidth=0.8, width=0.5)
        axes[0].set_title("Comparison", color="white")
        axes[0].set_ylabel("Amount ($)", color="white")
        axes[0].tick_params(colors="white")

        for bar, value in zip(bars, values):
            axes[0].text(
                bar.get_x() + bar.get_width()/2,
                bar.get_height() + max(values)*0.01,
                f"${value:.2f}",
                ha="center", va="bottom",
                fontsize=10, fontweight="bold",
                color="white")

        axes[1].pie(
            [max(income, 0.01), max(expenses, 0.01)],
            labels=["Income", "Expenses"],
            colors=bar_colors,
            autopct="%1.1f%%", startangle=90,
            wedgeprops={"edgecolor": "#1A2530",
                        "linewidth": 1.5},
            textprops={"color": "white"})

        balance_color = "#2ECC71" if balance >= 0 \
                        else "#E74C3C"
        balance_label = "Surplus" if balance >= 0 \
                        else "Deficit"
        axes[1].set_title("Split", color="white")
        axes[1].text(0, -1.35,
                     f"{balance_label}: ${abs(balance):.2f}",
                     ha="center", fontsize=12,
                     fontweight="bold",
                     color=balance_color)

        plt.tight_layout()
        plt.show()

    # ── Export ───────────────────────────────────────────────
    def _export_excel(self):
        if self.df is None:
            messagebox.showinfo("No Data",
                "Please load a CSV file first.")
            return

        filepath = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx")],
            title="Export to Excel"
        )
        if not filepath:
            return

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Transactions"

        headers     = ["Type", "Category",
                        "Amount", "Description"]
        header_fill = PatternFill("solid", fgColor="2B3A4A")
        header_font = Font(bold=True, color="FFFFFF", size=11)

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

        income_fill  = PatternFill("solid", fgColor="D5F5E3")
        expense_fill = PatternFill("solid", fgColor="FADBD8")

        for row, (_, t) in enumerate(
                self.df.iterrows(), 2):
            ws.cell(row=row, column=1,
                    value=str(t["type"]).capitalize())
            ws.cell(row=row, column=2,
                    value=t["category"])
            ws.cell(row=row, column=3,
                    value=round(float(t["amount"]), 2))
            ws.cell(row=row, column=4,
                    value=str(t.get("description", "")))
            fill = income_fill \
                   if t["type"] == "income" \
                   else expense_fill
            for col in range(1, 5):
                ws.cell(row=row, column=col).fill = fill

        income   = self.df[self.df["type"] == "income"]["amount"].sum()
        expenses = self.df[self.df["type"] == "expense"]["amount"].sum()
        balance  = income - expenses

        sr = len(self.df) + 3
        ws.cell(row=sr,   column=3,
                value="Total Income").font   = Font(bold=True)
        ws.cell(row=sr,   column=4,
                value=round(income, 2)).font = Font(
                bold=True, color="27AE60")
        ws.cell(row=sr+1, column=3,
                value="Total Expenses").font = Font(bold=True)
        ws.cell(row=sr+1, column=4,
                value=round(expenses, 2)).font = Font(
                bold=True, color="E74C3C")
        ws.cell(row=sr+2, column=3,
                value="Balance").font = Font(bold=True)
        ws.cell(row=sr+2, column=4,
                value=round(balance, 2)).font = Font(bold=True)

        ws.column_dimensions["A"].width = 12
        ws.column_dimensions["B"].width = 18
        ws.column_dimensions["C"].width = 12
        ws.column_dimensions["D"].width = 35

        wb.save(filepath)
        messagebox.showinfo("Exported",
            f"Data exported to:\n{filepath}")

    def _export_report(self):
        if self.df is None:
            messagebox.showinfo("No Data",
                "Please load a CSV file first.")
            return

        filepath = filedialog.asksaveasfilename(
            defaultextension=".txt",
            filetypes=[("Text files", "*.txt")],
            title="Save Report"
        )
        if not filepath:
            return

        income   = self.df[self.df["type"] == "income"]["amount"].sum()
        expenses = self.df[self.df["type"] == "expense"]["amount"].sum()
        balance  = income - expenses
        exp_df   = self.df[self.df["type"] == "expense"]
        breakdown = exp_df.groupby("category")["amount"] \
                          .sum().sort_values(ascending=False)
        top5     = exp_df.nlargest(5, "amount")
        timestamp = datetime.now().strftime("%Y-%m-%d %H:%M")

        with open(filepath, "w", encoding="utf-8") as f:
            f.write("=" * 44 + "\n")
            f.write("  CSV ANALYZER REPORT\n")
            f.write(f"  {timestamp}\n")
            f.write("=" * 44 + "\n\n")
            f.write(f"  Total Income:   ${income:.2f}\n")
            f.write(f"  Total Expenses: ${expenses:.2f}\n")
            f.write(f"  Balance:        ${balance:.2f}\n\n")
            f.write("  CATEGORY BREAKDOWN\n")
            f.write("  " + "-" * 30 + "\n")
            for cat, total in breakdown.items():
                f.write(f"  {cat:<16} ${total:.2f}\n")
            f.write("\n  TOP 5 EXPENSES\n")
            f.write("  " + "-" * 30 + "\n")
            for rank, (_, row) in enumerate(
                    top5.iterrows(), 1):
                f.write(f"  #{rank} ${row['amount']:.2f}"
                        f" — {row['category']}\n")

        messagebox.showinfo("Exported",
            f"Report saved to:\n{filepath}")

    # ── Dark Mode ────────────────────────────────────────────
    def _toggle_theme(self):
        self.dark_mode = not self.dark_mode
        self._apply_colors()

        if self.dark_mode:
            self.btn_theme.configure(text="☀️")
        else:
            self.btn_theme.configure(text="🌙")

        self.toolbar.configure(
            fg_color=self.C["slate_dark"]
            if self.dark_mode else self.C["slate"])

        # Update main background
        self.configure(fg_color=self.bg)

        # Update stats panel
        self.stats_panel.configure(fg_color=self.panel_bg)

        # Update all CTkFrames and CTkLabels recursively
        self._theme_recursive(self, self.panel_bg,
                               self.text_color)

        # Update Treeview style
        style = ttk.Style()
        style.configure("Analyst.Treeview",
                         background=self.panel_bg,
                         foreground=self.text_color,
                         fieldbackground=self.panel_bg)

        self.tree.tag_configure("income",
                                  background=self.income_row)
        self.tree.tag_configure("expense",
                                  background=self.expense_row)

        # Re-apply amber section headers
        for widget in self.stats_panel.winfo_children():
            if isinstance(widget, ctk.CTkLabel):
                text = widget.cget("text")
                if text in ("SUMMARY", "TOP CATEGORIES",
                            "STATISTICS"):
                    widget.configure(
                        text_color=self.C["amber"])
        # Refresh all stat row label colors
        for widget in self.stats_panel.winfo_children():
            try:
                for child in widget.winfo_children():
                    if isinstance(child, ctk.CTkLabel):
                        current = child.cget("text_color")
                        if current == "#6C757D" or \
                           current == "#8899AA" or \
                           current == "grey":
                            child.configure(
                                text_color=self.sub_color)
            except Exception:
                pass
        # Update all sub label colors
        for lbl in getattr(self, "_sub_labels", []):
            lbl.configure(text_color=self.sub_color)
        # Rebuild category labels with correct theme colors
        self._update_categories()   
        self.update_idletasks()
        self.update()

    def _theme_recursive(self, widget, panel_bg, text_color):
        try:
            wc = widget.winfo_class()
            if wc == "Frame":
                widget.configure(bg=panel_bg)
            elif wc in ("CTkFrame",):
                widget.configure(fg_color=panel_bg)
            elif wc == "CTkLabel":
                widget.configure(fg_color="transparent")
        except Exception:
            pass
        for child in widget.winfo_children():
            self._theme_recursive(child, panel_bg,
                                   text_color)

# ── Entry Point ──────────────────────────────────────────────
if __name__ == "__main__":
    app = CSVAnalyzerV11()
    app.mainloop()
