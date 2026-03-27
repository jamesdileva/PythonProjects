# budget_tracker_v11.py
# Budget Tracker v1.1 — CustomTkinter modern UI
# Navy and Emerald color scheme — two panel dashboard layout

import json
import os
import customtkinter as ctk
from tkinter import ttk, messagebox, filedialog
import matplotlib.pyplot as plt
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
import csv
from datetime import datetime

# ── CustomTkinter Setup ──────────────────────────────────────
ctk.set_appearance_mode("light")
ctk.set_default_color_theme("blue")

# ── Color Palette ────────────────────────────────────────────
COLORS = {
    "navy":        "#0D1B2A",
    "navy_light":  "#1B2E45",
    "emerald":     "#2ECC71",
    "emerald_dark":"#27AE60",
    "red":         "#E74C3C",
    "red_dark":    "#C0392B",
    "amber":       "#F39C12",
    "white":       "#FFFFFF",
    "off_white":   "#F8F9FA",
    "grey_light":  "#E9ECEF",
    "grey":        "#6C757D",
    "text_dark":   "#1A1A2E",
    "text_light":  "#F8F9FA",
    "income_bg":   "#D5F5E3",
    "expense_bg":  "#FADBD8",
    "dark_bg":     "#0D1B2A",
    "dark_panel":  "#1B2E45",
    "dark_surface":"#243B55",
}

SAVE_FILE = "transactions.json"

# ── Data Layer ───────────────────────────────────────────────
def load_transactions():
    if os.path.exists(SAVE_FILE):
        with open(SAVE_FILE, "r") as f:
            return json.load(f)
    return []

def save_transactions(transactions):
    with open(SAVE_FILE, "w") as f:
        json.dump(transactions, f, indent=2)

# ── Main Application ─────────────────────────────────────────
class BudgetTrackerV11(ctk.CTk):
    def __init__(self):
        super().__init__()

        self.title("Budget Tracker")
        self.geometry("1100x680")
        self.minsize(800, 600)

        self.transactions = load_transactions()
        self.dark_mode    = False

        self._apply_colors()
        self._build_ui()
        self._refresh()

    def _apply_colors(self):
        self.C = COLORS
        if self.dark_mode:
            ctk.set_appearance_mode("dark")
            self.bg         = self.C["dark_bg"]
            self.panel_bg   = self.C["dark_panel"]
            self.surface_bg = self.C["dark_surface"]
            self.text_color = self.C["text_light"]
            self.sub_color  = "#8899AA"
            self.income_row = "#1A3A2A"
            self.expense_row= "#3A1A1A"
        else:
            ctk.set_appearance_mode("light")
            self.bg         = self.C["off_white"]
            self.panel_bg   = self.C["white"]
            self.surface_bg = self.C["grey_light"]
            self.text_color = self.C["text_dark"]
            self.sub_color  = self.C["grey"]
            self.income_row = self.C["income_bg"]
            self.expense_row= self.C["expense_bg"]

    # ── UI Construction ──────────────────────────────────────
    def _build_ui(self):
        self.configure(fg_color=self.bg)
        self.columnconfigure(1, weight=1)
        self.rowconfigure(0, weight=1)

        self._build_left_panel()
        self._build_right_panel()

    def _build_left_panel(self):
        self.left = ctk.CTkFrame(self,
                                  fg_color=self.C["navy"],
                                  corner_radius=0,
                                  width=280)
        self.left.grid(row=0, column=0, sticky="nsew")
        self.left.grid_propagate(False)
        self.left.columnconfigure(0, weight=1)
        self.left.rowconfigure(0, weight=1)
        # Scrollable container for left panel content
        self.left_scroll = ctk.CTkScrollableFrame(
            self.left,
            fg_color=self.C["navy"],
            scrollbar_button_color=self.C["navy_light"],
            scrollbar_button_hover_color=self.C["emerald"]
        )
        self.left_scroll.grid(row=0, column=0, sticky="nsew")
        self.left_scroll.columnconfigure(0, weight=1)

        # Header
        ctk.CTkLabel(self.left_scroll,
                      text="💰 Budget Tracker",
                      font=ctk.CTkFont("Arial", 20, "bold"),
                      text_color=self.C["white"]).pack(
                      pady=(24,4), padx=20, anchor="w")

        ctk.CTkLabel(self.left_scroll,
                      text="Personal Finance Dashboard",
                      font=ctk.CTkFont("Arial", 11),
                      text_color="#8899AA").pack(
                      padx=20, anchor="w")

        ctk.CTkFrame(self.left_scroll, height=1,
                      fg_color="#1B2E45").pack(
                      fill="x", padx=20, pady=16)

        ctk.CTkLabel(self.left_scroll,
                      text="ADD TRANSACTION",
                      font=ctk.CTkFont("Arial", 10, "bold"),
                      text_color="#8899AA").pack(
                      padx=20, anchor="w", pady=(0,10))

        ctk.CTkLabel(self.left_scroll, text="Type",
                      font=ctk.CTkFont("Arial", 11),
                      text_color=self.C["text_light"]).pack(
                      padx=20, anchor="w")

        self.var_type = ctk.StringVar(value="expense")
        ctk.CTkOptionMenu(self.left_scroll,
                           variable=self.var_type,
                           values=["expense", "income"],
                           fg_color=self.C["navy_light"],
                           button_color=self.C["emerald"],
                           button_hover_color=self.C["emerald_dark"],
                           text_color=self.C["white"],
                           font=ctk.CTkFont("Arial", 11)).pack(
                           fill="x", padx=20, pady=(4,10))

        ctk.CTkLabel(self.left_scroll, text="Category",
                      font=ctk.CTkFont("Arial", 11),
                      text_color=self.C["text_light"]).pack(
                      padx=20, anchor="w")

        self.entry_cat = ctk.CTkEntry(self.left_scroll,
                                       placeholder_text="e.g. Salary, Rent, Food",
                                       fg_color=self.C["navy_light"],
                                       border_color=self.C["emerald"],
                                       text_color=self.C["white"],
                                       placeholder_text_color="#8899AA",
                                       font=ctk.CTkFont("Arial", 11))
        self.entry_cat.pack(fill="x", padx=20, pady=(4,10))

        ctk.CTkLabel(self.left_scroll, text="Amount ($)",
                      font=ctk.CTkFont("Arial", 11),
                      text_color=self.C["text_light"]).pack(
                      padx=20, anchor="w")

        self.entry_amt = ctk.CTkEntry(self.left_scroll,
                                       placeholder_text="0.00",
                                       fg_color=self.C["navy_light"],
                                       border_color=self.C["emerald"],
                                       text_color=self.C["white"],
                                       placeholder_text_color="#8899AA",
                                       font=ctk.CTkFont("Arial", 11))
        self.entry_amt.pack(fill="x", padx=20, pady=(4,10))

        ctk.CTkLabel(self.left_scroll, text="Description",
                      font=ctk.CTkFont("Arial", 11),
                      text_color=self.C["text_light"]).pack(
                      padx=20, anchor="w")

        self.entry_desc = ctk.CTkEntry(self.left_scroll,
                                        placeholder_text="What was this for?",
                                        fg_color=self.C["navy_light"],
                                        border_color=self.C["emerald"],
                                        text_color=self.C["white"],
                                        placeholder_text_color="#8899AA",
                                        font=ctk.CTkFont("Arial", 11))
        self.entry_desc.pack(fill="x", padx=20, pady=(4,16))

        ctk.CTkButton(self.left_scroll,
                       text="➕  Add Transaction",
                       command=self._add_transaction,
                       fg_color=self.C["emerald"],
                       hover_color=self.C["emerald_dark"],
                       text_color=self.C["white"],
                       font=ctk.CTkFont("Arial", 12, "bold"),
                       height=40,
                       corner_radius=8).pack(
                       fill="x", padx=20, pady=(0,6))

        ctk.CTkFrame(self.left_scroll, height=1,
                      fg_color="#1B2E45").pack(
                      fill="x", padx=20, pady=16)

        ctk.CTkLabel(self.left_scroll,
                      text="ACTIONS",
                      font=ctk.CTkFont("Arial", 10, "bold"),
                      text_color="#8899AA").pack(
                      padx=20, anchor="w", pady=(0,10))

        btn_cfg = {
            "font": ctk.CTkFont("Arial", 11),
            "height": 34,
            "corner_radius": 8,
            "anchor": "w"
        }

        ctk.CTkButton(self.left_scroll,
                       text="✏️  Edit Selected",
                       command=self._edit_transaction,
                       fg_color=self.C["amber"],
                       hover_color="#D68910",
                       text_color=self.C["white"],
                       **btn_cfg).pack(
                       fill="x", padx=20, pady=(0,6))

        ctk.CTkButton(self.left_scroll,
                       text="🗑  Delete Selected",
                       command=self._delete_transaction,
                       fg_color=self.C["red"],
                       hover_color=self.C["red_dark"],
                       text_color=self.C["white"],
                       **btn_cfg).pack(
                       fill="x", padx=20, pady=(0,6))

        ctk.CTkFrame(self.left_scroll, height=1,
                      fg_color="#1B2E45").pack(
                      fill="x", padx=20, pady=16)

        self.btn_theme = ctk.CTkButton(
                          self.left_scroll,
                          text="🌙  Dark Mode",
                          command=self._toggle_theme,
                          fg_color=self.C["navy_light"],
                          hover_color="#243B55",
                          text_color=self.C["white"],
                          font=ctk.CTkFont("Arial", 11),
                          height=34,
                          corner_radius=8)
        self.btn_theme.pack(fill="x", padx=20, pady=(0,20))

    def _build_right_panel(self):
        self.right = ctk.CTkFrame(self,
                                   fg_color=self.bg,
                                   corner_radius=0)
        self.right.grid(row=0, column=1, sticky="nsew")
        self.right.columnconfigure(0, weight=1)
        self.right.rowconfigure(1, weight=1)

        # Summary cards
        self._build_summary_cards()

        # Transaction table
        self._build_table()

        # Export buttons
        self._build_export_buttons()

    def _build_summary_cards(self):
        card_row = ctk.CTkFrame(self.right,
                                 fg_color="transparent")
        card_row.grid(row=0, column=0, sticky="ew",
                       padx=20, pady=(20,12))
        card_row.columnconfigure([0,1,2], weight=1)

        self.card_income   = self._make_card(
            card_row, "TOTAL INCOME", "$0.00",
            self.C["emerald"], 0)
        self.card_expenses = self._make_card(
            card_row, "TOTAL EXPENSES", "$0.00",
            self.C["red"], 1)
        self.card_balance  = self._make_card(
            card_row, "BALANCE", "$0.00",
            self.C["navy"], 2)

    def _make_card(self, parent, title, value, color, col):
        card = ctk.CTkFrame(parent,
                             fg_color=color,
                             corner_radius=12)
        card.grid(row=0, column=col, sticky="ew",
                  padx=(0 if col == 0 else 8, 0))

        ctk.CTkLabel(card,
                      text=title,
                      font=ctk.CTkFont("Arial", 10, "bold"),
                      text_color="#CCDDCC").pack(
                      padx=16, pady=(14,2), anchor="w")

        lbl = ctk.CTkLabel(card,
                            text=value,
                            font=ctk.CTkFont("Arial", 22, "bold"),
                            text_color=self.C["white"])
        lbl.pack(padx=16, pady=(0,14), anchor="w")

        return lbl

    def _build_table(self):
        table_frame = ctk.CTkFrame(self.right,
                                    fg_color=self.panel_bg,
                                    corner_radius=12)
        table_frame.grid(row=1, column=0, sticky="nsew",
                          padx=20, pady=(0,12))
        table_frame.columnconfigure(0, weight=1)
        table_frame.rowconfigure(1, weight=1)

        ctk.CTkLabel(table_frame,
                      text="Transactions",
                      font=ctk.CTkFont("Arial", 14, "bold"),
                      text_color=self.text_color).grid(
                      row=0, column=0, sticky="w",
                      padx=16, pady=(14,8))

        style = ttk.Style()
        style.theme_use("clam")
        style.configure("Modern.Treeview",
                         background=self.panel_bg,
                         foreground=self.text_color,
                         fieldbackground=self.panel_bg,
                         rowheight=32,
                         font=("Arial", 10))
        style.configure("Modern.Treeview.Heading",
                         background=self.C["navy"],
                         foreground=self.C["white"],
                         font=("Arial", 10, "bold"),
                         relief="flat")
        style.map("Modern.Treeview",
                   background=[("selected", self.C["navy_light"])])

        cols = ("type", "category", "amount", "description")
        self.tree = ttk.Treeview(table_frame,
                                  columns=cols,
                                  show="headings",
                                  style="Modern.Treeview",
                                  height=16)

        self.tree.heading("type",        text="Type")
        self.tree.heading("category",    text="Category")
        self.tree.heading("amount",      text="Amount")
        self.tree.heading("description", text="Description")

        self.tree.column("type",        width=90,  anchor="center")
        self.tree.column("category",    width=130, anchor="center")
        self.tree.column("amount",      width=100, anchor="center")
        self.tree.column("description", width=300, anchor="w")

        sb = ttk.Scrollbar(table_frame, orient="vertical",
                            command=self.tree.yview)
        self.tree.configure(yscrollcommand=sb.set)

        self.tree.grid(row=1, column=0, sticky="nsew",
                        padx=(12,0), pady=(0,12))
        sb.grid(row=1, column=1, sticky="ns", pady=(0,12))

    def _build_export_buttons(self):
        btn_row = ctk.CTkFrame(self.right,
                                fg_color="transparent")
        btn_row.grid(row=2, column=0, sticky="ew",
                      padx=20, pady=(0,20))

        export_cfg = {
            "font": ctk.CTkFont("Arial", 11),
            "height": 36,
            "corner_radius": 8,
            "width": 160
        }

        ctk.CTkButton(btn_row,
                       text="📊 Spending Chart",
                       command=self._show_chart,
                       fg_color="#8E44AD",
                       hover_color="#7D3C98",
                       text_color=self.C["white"],
                       **export_cfg).grid(
                       row=0, column=0, padx=(0,8), pady=(0,6))

        ctk.CTkButton(btn_row,
                       text="📈 Income vs Expenses",
                       command=self._show_income_chart,
                       fg_color="#16A085",
                       hover_color="#138D75",
                       text_color=self.C["white"],
                       **export_cfg).grid(
                       row=0, column=1, padx=(0,8), pady=(0,6))

        ctk.CTkButton(btn_row,
                       text="💾 Export Excel",
                       command=self._export_excel,
                       fg_color="#217346",
                       hover_color="#1A5C38",
                       text_color=self.C["white"],
                       **export_cfg).grid(
                       row=0, column=2, padx=(0,8), pady=(0,6))

        ctk.CTkButton(btn_row,
                       text="📄 Export CSV",
                       command=self._export_csv,
                       fg_color=self.C["amber"],
                       hover_color="#D68910",
                       text_color=self.C["white"],
                       **export_cfg).grid(
                       row=0, column=3, pady=(0,6))

    # ── Refresh ──────────────────────────────────────────────
    def _refresh(self):
        for row in self.tree.get_children():
            self.tree.delete(row)

        for t in self.transactions:
            tag = "income" if t["type"] == "income" else "expense"
            self.tree.insert("", "end", values=(
                t["type"].capitalize(),
                t["category"],
                f"${t['amount']:.2f}",
                t["description"]
            ), tags=(tag,))

        self.tree.tag_configure("income",
                                  background=self.income_row)
        self.tree.tag_configure("expense",
                                  background=self.expense_row)

        income   = sum(t["amount"] for t in self.transactions
                       if t["type"] == "income")
        expenses = sum(t["amount"] for t in self.transactions
                       if t["type"] == "expense")
        balance  = income - expenses

        self.card_income.configure(text=f"${income:.2f}")
        self.card_expenses.configure(text=f"${expenses:.2f}")
        self.card_balance.configure(text=f"${balance:.2f}")

    # ── Add Transaction ──────────────────────────────────────
    def _add_transaction(self):
        cat     = self.entry_cat.get().strip()
        amt_str = self.entry_amt.get().strip()
        desc    = self.entry_desc.get().strip()
        t_type  = self.var_type.get()

        if not cat or not desc or not amt_str:
            messagebox.showwarning("Missing Fields",
                "Please fill in all fields.")
            return

        if len(cat) < 2 or not any(c.isalpha() for c in cat):
            messagebox.showwarning("Invalid Category",
                "Category must be at least 2 letters.")
            return

        try:
            amount = float(amt_str)
            if amount <= 0:
                raise ValueError
        except ValueError:
            messagebox.showwarning("Invalid Amount",
                "Please enter a valid amount greater than zero.")
            return

        self.transactions.append({
            "type": t_type, "category": cat,
            "amount": amount, "description": desc
        })
        save_transactions(self.transactions)

        self.entry_cat.delete(0, "end")
        self.entry_amt.delete(0, "end")
        self.entry_desc.delete(0, "end")
        self._refresh()

    # ── Delete Transaction ───────────────────────────────────
    def _delete_transaction(self):
        selected = self.tree.selection()
        if not selected:
            messagebox.showwarning("Nothing Selected",
                "Please select a transaction to delete.")
            return
        if not messagebox.askyesno("Confirm",
                "Delete selected transaction?"):
            return
        index = self.tree.index(selected[0])
        self.transactions.pop(index)
        save_transactions(self.transactions)
        self._refresh()

    # ── Edit Transaction ─────────────────────────────────────
    def _edit_transaction(self):
        selected = self.tree.selection()
        if not selected:
            messagebox.showwarning("Nothing Selected",
                "Please select a transaction to edit.")
            return

        index = self.tree.index(selected[0])
        t = self.transactions[index]

        win = ctk.CTkToplevel(self)
        win.title("Edit Transaction")
        win.geometry("400x300")
        win.resizable(False, False)
        win.grab_set()
        win.configure(fg_color=self.C["navy"])

        ctk.CTkLabel(win, text="Edit Transaction",
                      font=ctk.CTkFont("Arial", 16, "bold"),
                      text_color=self.C["white"]).pack(
                      pady=(20,16), padx=20, anchor="w")

        var_type = ctk.StringVar(value=t["type"])
        ctk.CTkOptionMenu(win, variable=var_type,
                           values=["expense", "income"],
                           fg_color=self.C["navy_light"],
                           button_color=self.C["emerald"],
                           text_color=self.C["white"]).pack(
                           fill="x", padx=20, pady=(0,8))

        entry_cat = ctk.CTkEntry(win,
                                  fg_color=self.C["navy_light"],
                                  border_color=self.C["emerald"],
                                  text_color=self.C["white"],
                                  placeholder_text="Category")
        entry_cat.insert(0, t["category"])
        entry_cat.pack(fill="x", padx=20, pady=(0,8))

        entry_amt = ctk.CTkEntry(win,
                                  fg_color=self.C["navy_light"],
                                  border_color=self.C["emerald"],
                                  text_color=self.C["white"],
                                  placeholder_text="Amount")
        entry_amt.insert(0, str(t["amount"]))
        entry_amt.pack(fill="x", padx=20, pady=(0,8))

        entry_desc = ctk.CTkEntry(win,
                                   fg_color=self.C["navy_light"],
                                   border_color=self.C["emerald"],
                                   text_color=self.C["white"],
                                   placeholder_text="Description")
        entry_desc.insert(0, t["description"])
        entry_desc.pack(fill="x", padx=20, pady=(0,16))

        def save_edit():
            cat  = entry_cat.get().strip()
            desc = entry_desc.get().strip()
            try:
                amount = float(entry_amt.get().strip())
                if amount <= 0:
                    raise ValueError
            except ValueError:
                messagebox.showwarning("Invalid Amount",
                    "Please enter a valid amount.", parent=win)
                return
            if not cat or not desc:
                messagebox.showwarning("Missing Fields",
                    "Please fill in all fields.", parent=win)
                return
            self.transactions[index] = {
                "type": var_type.get(), "category": cat,
                "amount": amount, "description": desc
            }
            save_transactions(self.transactions)
            self._refresh()
            win.destroy()

        ctk.CTkButton(win, text="Save Changes",
                       command=save_edit,
                       fg_color=self.C["emerald"],
                       hover_color=self.C["emerald_dark"],
                       text_color=self.C["white"],
                       font=ctk.CTkFont("Arial", 12, "bold"),
                       height=38,
                       corner_radius=8).pack(
                       fill="x", padx=20)

    # ── Charts ───────────────────────────────────────────────
    def _show_chart(self):
        expenses = [t for t in self.transactions
                    if t["type"] == "expense"]
        if not expenses:
            messagebox.showinfo("No Data",
                "No expenses to chart yet.")
            return

        category_totals = {}
        for t in expenses:
            cat = t["category"]
            category_totals[cat] = \
                category_totals.get(cat, 0) + t["amount"]

        categories = list(category_totals.keys())
        amounts    = list(category_totals.values())
        colors     = ["#2ECC71","#E74C3C","#F39C12","#8E44AD",
                      "#16A085","#2E75B6","#D35400","#2C3E50"]

        fig, axes = plt.subplots(1, 2, figsize=(11, 5))
        fig.patch.set_facecolor("#0D1B2A")
        for ax in axes:
            ax.set_facecolor("#1B2E45")
            ax.tick_params(colors="white")
            ax.xaxis.label.set_color("white")
            ax.yaxis.label.set_color("white")
            ax.title.set_color("white")
            for spine in ax.spines.values():
                spine.set_edgecolor("#243B55")

        fig.suptitle("Spending by Category",
                     fontsize=14, fontweight="bold",
                     color="white")

        bars = axes[0].bar(categories, amounts,
                           color=colors[:len(categories)],
                           edgecolor="#0D1B2A", linewidth=0.8)
        axes[0].set_title("Expenses by Category")
        axes[0].set_xlabel("Category")
        axes[0].set_ylabel("Amount ($)")
        axes[0].tick_params(axis="x", rotation=30)

        for bar, amount in zip(bars, amounts):
            axes[0].text(
                bar.get_x() + bar.get_width()/2,
                bar.get_height() + max(amounts)*0.01,
                f"${amount:.0f}",
                ha="center", va="bottom",
                fontsize=9, color="white")

        axes[1].pie(amounts, labels=categories,
                    colors=colors[:len(categories)],
                    autopct="%1.1f%%", startangle=140,
                    wedgeprops={"edgecolor": "#0D1B2A",
                                "linewidth": 1.5},
                    textprops={"color": "white"})
        axes[1].set_title("Spending Distribution")

        plt.tight_layout()
        plt.show()

    def _show_income_chart(self):
        income   = sum(t["amount"] for t in self.transactions
                       if t["type"] == "income")
        expenses = sum(t["amount"] for t in self.transactions
                       if t["type"] == "expense")
        balance  = income - expenses

        if income == 0 and expenses == 0:
            messagebox.showinfo("No Data",
                "No transactions to chart yet.")
            return

        fig, axes = plt.subplots(1, 2, figsize=(11, 5))
        fig.patch.set_facecolor("#0D1B2A")
        for ax in axes:
            ax.set_facecolor("#1B2E45")
            ax.tick_params(colors="white")
            for spine in ax.spines.values():
                spine.set_edgecolor("#243B55")

        fig.suptitle("Income vs Expenses",
                     fontsize=14, fontweight="bold",
                     color="white")

        labels = ["Income", "Expenses"]
        values = [income, expenses]
        bar_colors = ["#2ECC71", "#E74C3C"]

        bars = axes[0].bar(labels, values,
                           color=bar_colors,
                           edgecolor="#0D1B2A",
                           linewidth=0.8, width=0.5)
        axes[0].set_title("Comparison",
                           color="white")
        axes[0].set_ylabel("Amount ($)", color="white")
        axes[0].tick_params(colors="white")

        for bar, value in zip(bars, values):
            axes[0].text(
                bar.get_x() + bar.get_width()/2,
                bar.get_height() + max(values)*0.01,
                f"${value:.2f}",
                ha="center", va="bottom",
                fontsize=10, fontweight="bold",
                color="white")

        axes[1].pie(
            [max(income, 0.01), max(expenses, 0.01)],
            labels=["Income", "Expenses"],
            colors=bar_colors,
            autopct="%1.1f%%", startangle=90,
            wedgeprops={"edgecolor": "#0D1B2A",
                        "linewidth": 1.5},
            textprops={"color": "white"})

        balance_color  = "#2ECC71" if balance >= 0 else "#E74C3C"
        balance_label  = "Surplus" if balance >= 0 else "Deficit"
        axes[1].set_title("Split", color="white")
        axes[1].text(0, -1.35,
                     f"{balance_label}: ${abs(balance):.2f}",
                     ha="center", fontsize=12,
                     fontweight="bold", color=balance_color)

        plt.tight_layout()
        plt.show()

    # ── Export ───────────────────────────────────────────────
    def _export_excel(self):
        if not self.transactions:
            messagebox.showinfo("No Data",
                "No transactions to export.")
            return

        filepath = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx")],
            title="Export to Excel"
        )
        if not filepath:
            return

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Transactions"

        headers     = ["Type", "Category", "Amount", "Description"]
        header_fill = PatternFill("solid", fgColor="0D1B2A")
        header_font = Font(bold=True, color="FFFFFF", size=11)

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

        income_fill  = PatternFill("solid", fgColor="D5F5E3")
        expense_fill = PatternFill("solid", fgColor="FADBD8")

        for row, t in enumerate(self.transactions, 2):
            ws.cell(row=row, column=1,
                    value=t["type"].capitalize())
            ws.cell(row=row, column=2, value=t["category"])
            ws.cell(row=row, column=3,
                    value=round(t["amount"], 2))
            ws.cell(row=row, column=4,
                    value=t["description"])
            fill = income_fill if t["type"] == "income" \
                   else expense_fill
            for col in range(1, 5):
                ws.cell(row=row, column=col).fill = fill

        income   = sum(t["amount"] for t in self.transactions
                       if t["type"] == "income")
        expenses = sum(t["amount"] for t in self.transactions
                       if t["type"] == "expense")
        balance  = income - expenses

        sr = len(self.transactions) + 3
        ws.cell(row=sr,   column=3,
                value="Total Income").font   = Font(bold=True)
        ws.cell(row=sr,   column=4,
                value=round(income, 2)).font = Font(bold=True,
                color="27AE60")
        ws.cell(row=sr+1, column=3,
                value="Total Expenses").font = Font(bold=True)
        ws.cell(row=sr+1, column=4,
                value=round(expenses, 2)).font = Font(bold=True,
                color="E74C3C")
        ws.cell(row=sr+2, column=3,
                value="Balance").font = Font(bold=True)
        ws.cell(row=sr+2, column=4,
                value=round(balance, 2)).font = Font(bold=True)

        ws.column_dimensions["A"].width = 12
        ws.column_dimensions["B"].width = 16
        ws.column_dimensions["C"].width = 12
        ws.column_dimensions["D"].width = 30

        wb.save(filepath)
        messagebox.showinfo("Exported",
            f"Transactions exported to:\n{filepath}")

    def _export_csv(self):
        if not self.transactions:
            messagebox.showinfo("No Data",
                "No transactions to export.")
            return

        filepath = filedialog.asksaveasfilename(
            defaultextension=".csv",
            filetypes=[("CSV files", "*.csv")],
            title="Export to CSV"
        )
        if not filepath:
            return

        with open(filepath, "w", newline="",
                  encoding="utf-8") as f:
            writer = csv.DictWriter(f,
                fieldnames=["type", "category",
                            "amount", "description"])
            writer.writeheader()
            writer.writerows(self.transactions)

        messagebox.showinfo("Exported",
            f"Transactions exported to:\n{filepath}")

    # ── Dark Mode ────────────────────────────────────────────
    def _toggle_theme(self):
        self.dark_mode = not self.dark_mode
        self._apply_colors()

        if self.dark_mode:
            ctk.set_appearance_mode("dark")
            self.btn_theme.configure(text="☀️  Light Mode")
        else:
            ctk.set_appearance_mode("light")
            self.btn_theme.configure(text="🌙  Dark Mode")

        self.tree.tag_configure("income",
                                  background=self.income_row)
        self.tree.tag_configure("expense",
                                  background=self.expense_row)

        style = ttk.Style()
        style.configure("Modern.Treeview",
                         background=self.panel_bg,
                         foreground=self.text_color,
                         fieldbackground=self.panel_bg)
        # Force full redraw
        self.right.configure(fg_color=self.bg)
        self.update_idletasks()
        self.update()
# ── Entry Point ──────────────────────────────────────────────
if __name__ == "__main__":
    app = BudgetTrackerV11()
    app.mainloop()
