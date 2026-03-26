# budget_tracker_gui.py
# Budget Tracker with GUI — built with Tkinter and Matplotlib

import json
import os
import tkinter as tk
from tkinter import ttk, messagebox
import matplotlib.pyplot as plt
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
from tkinter import filedialog
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

THEMES = {
    "light": {
        "bg": "#F0F0F0", "fg": "#1A1A1A", "summary_bg": "#2E75B6",
        "summary_fg": "white", "form_bg": "#F0F0F0", "btn_bg": "#F0F0F0",
        "tree_bg": "white", "income_bg": "#EAFAF1", "expense_bg": "#FDEDEC"
    },
    "dark": {
        "bg": "#1E1E1E", "fg": "#F0F0F0", "summary_bg": "#1A3A5C",
        "summary_fg": "#F0F0F0", "form_bg": "#2D2D2D", "btn_bg": "#2D2D2D",
        "tree_bg": "#2D2D2D", "income_bg": "#1E3A2F", "expense_bg": "#3A1E1E"
    }
}

SAVE_FILE = "transactions.json"

# ── Data Layer ──────────────────────────────────────────────
def load_transactions():
    if os.path.exists(SAVE_FILE):
        with open(SAVE_FILE, "r") as f:
            return json.load(f)
    return []

def save_transactions(transactions):
    with open(SAVE_FILE, "w") as f:
        json.dump(transactions, f, indent=2)

# ── Main Application Class ───────────────────────────────────
class BudgetApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Budget Tracker")
        self.root.geometry("900x650")
        self.root.resizable(True, True)
        self.transactions = load_transactions()
        self.theme = "light"
        self._build_ui()
        self._refresh()

    def _build_ui(self):
        self.root.columnconfigure(0, weight=1)
        self.root.rowconfigure(1, weight=1)

        # ── Top summary bar ──
        summary_frame = tk.Frame(self.root, bg="#2E75B6", pady=12)
        summary_frame.grid(row=0, column=0, sticky="ew")
        summary_frame.columnconfigure([0,1,2], weight=1)

        self.lbl_income = tk.Label(summary_frame, text="Income: $0.00",
            font=("Arial", 13, "bold"), bg="#2E75B6", fg="white")
        self.lbl_income.grid(row=0, column=0)

        self.lbl_expenses = tk.Label(summary_frame, text="Expenses: $0.00",
            font=("Arial", 13, "bold"), bg="#2E75B6", fg="white")
        self.lbl_expenses.grid(row=0, column=1)

        self.lbl_balance = tk.Label(summary_frame, text="Balance: $0.00",
            font=("Arial", 13, "bold"), bg="#2E75B6", fg="white")
        self.lbl_balance.grid(row=0, column=2)

        # ── Middle — transactions list ──
        middle_frame = tk.Frame(self.root)
        middle_frame.grid(row=1, column=0, sticky="nsew", padx=16, pady=10)
        middle_frame.columnconfigure(0, weight=1)
        middle_frame.rowconfigure(0, weight=1)

        columns = ("type", "category", "amount", "description")
        self.tree = ttk.Treeview(middle_frame, columns=columns,
                                  show="headings", height=18)

        self.tree.heading("type",        text="Type")
        self.tree.heading("category",    text="Category")
        self.tree.heading("amount",      text="Amount")
        self.tree.heading("description", text="Description")

        self.tree.column("type",        width=90,  anchor="center")
        self.tree.column("category",    width=130, anchor="center")
        self.tree.column("amount",      width=100, anchor="center")
        self.tree.column("description", width=300, anchor="w")

        scrollbar = ttk.Scrollbar(middle_frame, orient="vertical",
                                   command=self.tree.yview)
        self.tree.configure(yscrollcommand=scrollbar.set)

        self.tree.grid(row=0, column=0, sticky="nsew")
        scrollbar.grid(row=0, column=1, sticky="ns")

        # ── Bottom — input form ──
        form_frame = tk.LabelFrame(self.root, text="Add Transaction",
                                    font=("Arial", 10, "bold"), padx=12, pady=8)
        form_frame.grid(row=2, column=0, sticky="ew", padx=16, pady=(0, 8))

        tk.Label(form_frame, text="Type:",        font=("Arial", 10)).grid(row=0, column=0, sticky="w")
        tk.Label(form_frame, text="Category:",    font=("Arial", 10)).grid(row=0, column=2, sticky="w", padx=(16,0))
        tk.Label(form_frame, text="Amount ($):",  font=("Arial", 10)).grid(row=0, column=4, sticky="w", padx=(16,0))
        tk.Label(form_frame, text="Description:", font=("Arial", 10)).grid(row=0, column=6, sticky="w", padx=(16,0))

        self.var_type = tk.StringVar(value="expense")
        type_menu = ttk.Combobox(form_frame, textvariable=self.var_type,
                                  values=["income", "expense"], width=10, state="readonly")
        type_menu.grid(row=0, column=1, padx=(4,0))

        self.var_category = tk.StringVar()
        tk.Entry(form_frame, textvariable=self.var_category, width=14).grid(row=0, column=3, padx=(4,0))

        self.var_amount = tk.StringVar()
        tk.Entry(form_frame, textvariable=self.var_amount, width=10).grid(row=0, column=5, padx=(4,0))

        self.var_description = tk.StringVar()
        tk.Entry(form_frame, textvariable=self.var_description, width=22).grid(row=0, column=7, padx=(4,0))

        # ── Buttons ──
        self.btn_frame = tk.Frame(self.root)
        self.btn_frame.grid(row=3, column=0, pady=(0, 12))

        btn_style = {"font": ("Arial", 10, "bold"), "width": 16,
                     "pady": 6, "cursor": "hand2"}

        tk.Button(self.btn_frame, text="➕ Add Transaction",
                  command=self._add_transaction,
                  bg="#27AE60", fg="white", **btn_style).grid(row=0, column=0, padx=6)

        tk.Button(self.btn_frame, text="🗑 Delete Selected",
                  command=self._delete_transaction,
                  bg="#E74C3C", fg="white", **btn_style).grid(row=0, column=1, padx=6)

        tk.Button(self.btn_frame, text="✏️ Edit Selected",
                  command=self._edit_transaction,
                  bg="#F39C12", fg="white", **btn_style).grid(row=0, column=2, padx=6)

        tk.Button(self.btn_frame, text="📊 Spending Chart",
                  command=self._show_chart,
                  bg="#8E44AD", fg="white", **btn_style).grid(row=0, column=3, padx=6)

        tk.Button(self.btn_frame, text="📈 Income vs Expenses",
                  command=self._show_income_vs_expenses,
                  bg="#16A085", fg="white", **btn_style).grid(row=1, column=0, padx=6, pady=(6,0))

        tk.Button(self.btn_frame, text="💾 Export to Excel",
                  command=self._export_excel,
                  bg="#217346", fg="white", **btn_style).grid(row=1, column=1, padx=6, pady=(6,0))

        tk.Button(self.btn_frame, text="📄 Export to CSV",
                  command=self._export_csv,
                  bg="#E67E22", fg="white", **btn_style).grid(row=1, column=2, padx=6, pady=(6,0))

        tk.Button(self.btn_frame, text="🌙 Toggle Dark Mode",
                  command=self._toggle_theme,
                  bg="#2C3E50", fg="white", **btn_style).grid(row=1, column=3, padx=6, pady=(6,0))

    # ── Logic Methods ────────────────────────────────────────
    def _refresh(self):
        for row in self.tree.get_children():
            self.tree.delete(row)

        for t in self.transactions:
            tag = "income" if t["type"] == "income" else "expense"
            self.tree.insert("", "end", values=(
                t["type"].capitalize(),
                t["category"],
                f"${t['amount']:.2f}",
                t["description"]
            ), tags=(tag,))

        self.tree.tag_configure("income",  background="#EAFAF1")
        self.tree.tag_configure("expense", background="#FDEDEC")

        income   = sum(t["amount"] for t in self.transactions if t["type"] == "income")
        expenses = sum(t["amount"] for t in self.transactions if t["type"] == "expense")
        balance  = income - expenses

        self.lbl_income.config(text=f"Income: ${income:.2f}")
        self.lbl_expenses.config(text=f"Expenses: ${expenses:.2f}")
        self.lbl_balance.config(text=f"Balance: ${balance:.2f}",
            fg="white" if balance >= 0 else "#FFD700")

    def _add_transaction(self):
        t_type  = self.var_type.get().strip()
        cat     = self.var_category.get().strip()
        desc    = self.var_description.get().strip()
        amt_str = self.var_amount.get().strip()

        if not cat or not desc or not amt_str:
            messagebox.showwarning("Missing Fields", "Please fill in all fields.")
            return

        if len(cat) < 2 or not any(c.isalpha() for c in cat):
            messagebox.showwarning("Invalid Category",
                "Category must be at least 2 characters and contain letters.")
            return

        try:
            amount = float(amt_str)
            if amount <= 0:
                raise ValueError
        except ValueError:
            messagebox.showwarning("Invalid Amount",
                "Please enter a valid amount greater than zero.")
            return

        self.transactions.append({
            "type": t_type, "category": cat,
            "amount": amount, "description": desc
        })
        save_transactions(self.transactions)

        self.var_category.set("")
        self.var_amount.set("")
        self.var_description.set("")
        self._refresh()

    def _delete_transaction(self):
        selected = self.tree.selection()
        if not selected:
            messagebox.showwarning("Nothing Selected",
                "Please select a transaction to delete.")
            return

        if not messagebox.askyesno("Confirm Delete",
                "Delete the selected transaction?"):
            return

        index = self.tree.index(selected[0])
        self.transactions.pop(index)
        save_transactions(self.transactions)
        self._refresh()

    def _edit_transaction(self):
        selected = self.tree.selection()
        if not selected:
            messagebox.showwarning("Nothing Selected",
                "Please select a transaction to edit.")
            return

        index = self.tree.index(selected[0])
        t = self.transactions[index]

        edit_win = tk.Toplevel(self.root)
        edit_win.title("Edit Transaction")
        edit_win.geometry("400x220")
        edit_win.resizable(False, False)
        edit_win.grab_set()

        tk.Label(edit_win, text="Type:",        font=("Arial", 10)).grid(row=0, column=0, padx=12, pady=8, sticky="w")
        tk.Label(edit_win, text="Category:",    font=("Arial", 10)).grid(row=1, column=0, padx=12, pady=8, sticky="w")
        tk.Label(edit_win, text="Amount ($):",  font=("Arial", 10)).grid(row=2, column=0, padx=12, pady=8, sticky="w")
        tk.Label(edit_win, text="Description:", font=("Arial", 10)).grid(row=3, column=0, padx=12, pady=8, sticky="w")

        var_type = tk.StringVar(value=t["type"])
        var_cat  = tk.StringVar(value=t["category"])
        var_amt  = tk.StringVar(value=str(t["amount"]))
        var_desc = tk.StringVar(value=t["description"])

        ttk.Combobox(edit_win, textvariable=var_type,
                     values=["income", "expense"],
                     width=18, state="readonly").grid(row=0, column=1, padx=12)
        tk.Entry(edit_win, textvariable=var_cat,  width=20).grid(row=1, column=1, padx=12)
        tk.Entry(edit_win, textvariable=var_amt,  width=20).grid(row=2, column=1, padx=12)
        tk.Entry(edit_win, textvariable=var_desc, width=20).grid(row=3, column=1, padx=12)

        def save_edit():
            cat  = var_cat.get().strip()
            desc = var_desc.get().strip()
            amt_str = var_amt.get().strip()

            if not cat or not desc or not amt_str:
                messagebox.showwarning("Missing Fields",
                    "Please fill in all fields.", parent=edit_win)
                return
            try:
                amount = float(amt_str)
                if amount <= 0:
                    raise ValueError
            except ValueError:
                messagebox.showwarning("Invalid Amount",
                    "Please enter a valid amount greater than zero.",
                    parent=edit_win)
                return

            self.transactions[index] = {
                "type": var_type.get(),
                "category": cat,
                "amount": amount,
                "description": desc
            }
            save_transactions(self.transactions)
            self._refresh()
            edit_win.destroy()

        tk.Button(edit_win, text="Save Changes",
                  command=save_edit, bg="#27AE60", fg="white",
                  font=("Arial", 10, "bold"), width=14,
                  pady=6).grid(row=4, column=0, columnspan=2, pady=12)

    def _export_excel(self):
        if not self.transactions:
            messagebox.showinfo("No Data", "No transactions to export.")
            return

        filepath = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx")],
            title="Save Excel File"
        )
        if not filepath:
            return

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Transactions"

        # Header row
        headers = ["Type", "Category", "Amount", "Description"]
        header_fill = PatternFill("solid", fgColor="2E75B6")
        header_font = Font(bold=True, color="FFFFFF", size=11)

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

        # Data rows
        income_fill  = PatternFill("solid", fgColor="EAFAF1")
        expense_fill = PatternFill("solid", fgColor="FDEDEC")

        for row, t in enumerate(self.transactions, 2):
            ws.cell(row=row, column=1, value=t["type"].capitalize())
            ws.cell(row=row, column=2, value=t["category"])
            ws.cell(row=row, column=3, value=round(t["amount"], 2))
            ws.cell(row=row, column=4, value=t["description"])

            fill = income_fill if t["type"] == "income" else expense_fill
            for col in range(1, 5):
                ws.cell(row=row, column=col).fill = fill

        # Summary rows
        income   = sum(t["amount"] for t in self.transactions if t["type"] == "income")
        expenses = sum(t["amount"] for t in self.transactions if t["type"] == "expense")
        balance  = income - expenses

        summary_row = len(self.transactions) + 3
        ws.cell(row=summary_row,     column=3, value="Total Income").font   = Font(bold=True)
        ws.cell(row=summary_row,     column=4, value=round(income, 2)).font = Font(bold=True, color="27AE60")
        ws.cell(row=summary_row + 1, column=3, value="Total Expenses").font   = Font(bold=True)
        ws.cell(row=summary_row + 1, column=4, value=round(expenses, 2)).font = Font(bold=True, color="E74C3C")
        ws.cell(row=summary_row + 2, column=3, value="Balance").font   = Font(bold=True)
        ws.cell(row=summary_row + 2, column=4, value=round(balance, 2)).font = Font(bold=True)

        # Column widths
        ws.column_dimensions["A"].width = 12
        ws.column_dimensions["B"].width = 16
        ws.column_dimensions["C"].width = 12
        ws.column_dimensions["D"].width = 30

        wb.save(filepath)
        messagebox.showinfo("Exported",
            f"Transactions exported to:\n{filepath}")

    def _export_csv(self):
        if not self.transactions:
            messagebox.showinfo("No Data", "No transactions to export.")
            return

        filepath = filedialog.asksaveasfilename(
            defaultextension=".csv",
            filetypes=[("CSV files", "*.csv")],
            title="Save CSV File"
        )
        if not filepath:
            return

        import csv
        with open(filepath, "w", newline="", encoding="utf-8") as f:
            writer = csv.DictWriter(f,
                fieldnames=["type", "category", "amount", "description"])
            writer.writeheader()
            writer.writerows(self.transactions)

        messagebox.showinfo("Exported",
            f"Transactions exported to:\n{filepath}")

    def _toggle_theme(self):
        self.theme = "dark" if self.theme == "light" else "light"
        t = THEMES[self.theme]

        self.root.configure(bg=t["bg"])
        self._apply_theme_recursive(self.root, t)
        self._refresh()

    def _apply_theme_recursive(self, widget, t):
        widget_type = widget.winfo_class()
        try:
            if widget_type in ("Frame", "Labelframe", "Label"):
                widget.configure(bg=t["bg"], fg=t["fg"])
            elif widget_type == "Button":
                pass
        except tk.TclError:
            pass
        for child in widget.winfo_children():
            self._apply_theme_recursive(child, t)

    def _show_chart(self):
        expenses = [t for t in self.transactions if t["type"] == "expense"]
        if not expenses:
            messagebox.showinfo("No Data", "No expenses to chart yet.")
            return

        category_totals = {}
        for t in expenses:
            cat = t["category"]
            category_totals[cat] = category_totals.get(cat, 0) + t["amount"]

        categories = list(category_totals.keys())
        amounts    = list(category_totals.values())
        colors     = ["#2E75B6","#27AE60","#E74C3C","#F39C12",
                      "#8E44AD","#16A085","#D35400","#2C3E50"]

        fig, axes = plt.subplots(1, 2, figsize=(11, 5))
        fig.suptitle("Spending by Category", fontsize=14, fontweight="bold")

        # Bar chart
        bars = axes[0].bar(categories, amounts,
                           color=colors[:len(categories)], edgecolor="white", linewidth=0.8)
        axes[0].set_title("Expenses by Category")
        axes[0].set_xlabel("Category")
        axes[0].set_ylabel("Amount ($)")
        axes[0].tick_params(axis="x", rotation=30)
        for bar, amount in zip(bars, amounts):
            axes[0].text(bar.get_x() + bar.get_width()/2,
                         bar.get_height() + max(amounts)*0.01,
                         f"${amount:.0f}", ha="center", va="bottom", fontsize=9)

        # Pie chart
        axes[1].pie(amounts, labels=categories,
                    colors=colors[:len(categories)],
                    autopct="%1.1f%%", startangle=140,
                    wedgeprops={"edgecolor": "white", "linewidth": 1.5})
        axes[1].set_title("Spending Distribution")

        plt.tight_layout()
        plt.show()

    def _show_income_vs_expenses(self):
        income   = sum(t["amount"] for t in self.transactions if t["type"] == "income")
        expenses = sum(t["amount"] for t in self.transactions if t["type"] == "expense")
        balance  = income - expenses

        if income == 0 and expenses == 0:
            messagebox.showinfo("No Data", "No transactions to chart yet.")
            return

        fig, axes = plt.subplots(1, 2, figsize=(11, 5))
        fig.suptitle("Income vs Expenses Overview", fontsize=14, fontweight="bold")

        # Bar comparison
        labels = ["Income", "Expenses"]
        values = [income, expenses]
        bar_colors = ["#27AE60", "#E74C3C"]
        bars = axes[0].bar(labels, values, color=bar_colors,
                           edgecolor="white", linewidth=0.8, width=0.5)
        axes[0].set_title("Income vs Expenses")
        axes[0].set_ylabel("Amount ($)")
        for bar, value in zip(bars, values):
            axes[0].text(bar.get_x() + bar.get_width()/2,
                         bar.get_height() + max(values)*0.01,
                         f"${value:.2f}", ha="center", va="bottom",
                         fontsize=10, fontweight="bold")

        # Balance indicator
        balance_color = "#27AE60" if balance >= 0 else "#E74C3C"
        balance_label = "Surplus" if balance >= 0 else "Deficit"
        axes[1].pie(
            [max(income, 0.01), max(expenses, 0.01)],
            labels=["Income", "Expenses"],
            colors=["#27AE60", "#E74C3C"],
            autopct="%1.1f%%", startangle=90,
            wedgeprops={"edgecolor": "white", "linewidth": 1.5}
        )
        axes[1].set_title("Income vs Expenses Split")
        axes[1].text(0, -1.35,
                     f"{balance_label}: ${abs(balance):.2f}",
                     ha="center", fontsize=12, fontweight="bold",
                     color=balance_color)

        plt.tight_layout()
        plt.show()

# ── Entry Point ──────────────────────────────────────────────
if __name__ == "__main__":
    root = tk.Tk()
    app = BudgetApp(root)
    root.mainloop()
